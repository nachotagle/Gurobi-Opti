

from typing import Optional

def read_params_from_excel(filename: str = "parametros_reales.xlsx", sheet_names=None):
	"""Read parameters from an Excel file into a dictionary.

	The sheet should contain two columns: key | value
	Values can be numbers, strings, or Python literals (dicts, lists) encoded as text.
	The function will attempt to parse string values using ast.literal_eval, falling back
	to the raw string if parsing fails.

	Parameters
	- filename: path to the xlsx file
	- sheet_names: optional list of sheet names to try (in order). If None, will try
	  ['parametros_reales', 'params'] then the active sheet.
	"""
	try:
		from openpyxl import load_workbook
	except Exception:
		print("openpyxl is required to read Excel files. Install with: pip install openpyxl")
		return {}

	import ast

	try:
		wb = load_workbook(filename=filename, data_only=True)
	except Exception as e:
		print(f"Could not open Excel file {filename}: {e}")
		return {}

	if sheet_names is None:
		sheet_names = ["parametros_reales", "params"]

	ws = None
	for s in sheet_names:
		if s in wb.sheetnames:
			ws = wb[s]
			break

	if ws is None:
		# fallback to active sheet
		ws = wb.active

	params = {}
	for row in ws.iter_rows(min_row=1, values_only=True):
		if not row or row[0] is None:
			continue
		key = str(row[0]).strip()
		val_cell = row[1] if len(row) > 1 else None
		if val_cell is None:
			params[key] = None
			continue

		# attempt to parse
		if isinstance(val_cell, str):
			s = val_cell.strip()
			try:
				parsed = ast.literal_eval(s)
			except Exception:
				parsed = s
			params[key] = parsed
		else:
			# numeric, boolean, date, etc.
			params[key] = val_cell

	return params


def write_solution_to_excel(m, filename: str = "solution.xlsx", include_zeros: bool = False, sheet_name: str = "Solution"):
	try:
		import xlsxwriter
	except Exception:
		print("XlsxWriter is required to write Excel files. Install with: pip install XlsxWriter")
		return

	# parse variables into (base, indices) map
	parsed = {}  # base -> dict of tuple(indices) -> value
	max_day = 0

	for v in m.getVars():
		name = v.VarName
		if "[" in name and name.endswith("]"):
			base, idxs = name.split("[", 1)
			idxs = idxs[:-1]
			parts = [p.strip() for p in idxs.split(",")]
			indices = []
			for p in parts:
				try:
					indices.append(int(p))
				except Exception:
					indices.append(p)
			key = tuple(indices)
		else:
			base = name
			key = tuple()

		# get value
		val = None
		try:
			val = v.X
		except Exception:
			try:
				val = v.getAttr("X")
			except Exception:
				val = None

		if val is None:
			continue
		if not include_zeros and abs(val) < 1e-12:
			continue

		parsed.setdefault(base, {})[key] = val
		# if last index numeric, consider it a day
		if len(key) >= 1 and isinstance(key[-1], int):
			if key[-1] > max_day:
				max_day = key[-1]

	if max_day == 0:
		# no day-indexed variables found; fall back to previous behavior: list variables vertically
		try:
			wb = xlsxwriter.Workbook(filename)
			ws = wb.add_worksheet(sheet_name)
		except Exception as e:
			print("Could not create Excel workbook:", e)
			return

		# write objective and status in top-left summary
		status = getattr(m, "status", None)
		try:
			obj = m.objVal
		except Exception:
			try:
				obj = m.getAttr("ObjVal")
			except Exception:
				obj = None

		ws.write(0, 0, "Z*")
		ws.write(0, 1, obj if obj is not None else "(n/a)")
		ws.write(1, 0, "Model status:")
		ws.write(1, 1, str(status))

		start_row = 3
		# header
		ws.write(start_row, 0, "Variable")
		for j, base in enumerate(sorted(parsed.keys()), start=1):
			ws.write(start_row, j, base)

		# single row with values (use first key if multiple)
		for j, base in enumerate(sorted(parsed.keys()), start=1):
			vals = parsed[base]
			# pick a representative value
			if vals:
				first = next(iter(vals.values()))
				ws.write(start_row + 1, j, first)

		try:
			wb.close()
			print(f"Wrote solution to {filename}")
		except Exception as e:
			print("Could not save Excel file:", e)
		return

	# Build column list: for each base, for each prefix (indices except last) create a column
	columns = []  # list of (base, prefix_key)
	for base, table in parsed.items():
		# collect all keys that end with numeric day
		prefixes = set()
		has_day = False
		for k in table.keys():
			if len(k) >= 1 and isinstance(k[-1], int):
				has_day = True
				prefixes.add(k[:-1])
			else:
				# scalar or non-day-indexed -> treat as prefix ()
				prefixes.add(k)
		if has_day:
			for p in sorted(prefixes):
				columns.append((base, p))

	# create workbook and sheet
	try:
		wb = xlsxwriter.Workbook(filename)
		ws = wb.add_worksheet(sheet_name)
	except Exception as e:
		print("Could not create Excel workbook:", e)
		return

	# write objective and status in top-left summary
	status = getattr(m, "status", None)
	try:
		obj = m.objVal
	except Exception:
		try:
			obj = m.getAttr("ObjVal")
		except Exception:
			obj = None

	ws.write(0, 0, "Z*")
	ws.write(0, 1, obj if obj is not None else "(n/a)")
	ws.write(1, 0, "Model status:")
	ws.write(1, 1, str(status))

	start_row = 3
	# Header row: Day then one column per variable-instance
	ws.write(start_row, 0, "Day")
	for j, (base, pk) in enumerate(columns, start=1):
		if pk == ():
			colname = base
		else:
			colname = base + "_" + "_".join(str(x) for x in pk)
		ws.write(start_row, j, colname)

	# Fill rows for days 1..max_day
	for day in range(1, max_day + 1):
		ws.write(start_row + day, 0, day)
		for j, (base, pk) in enumerate(columns, start=1):
			key = pk + (day,) if pk != () else (day,)
			val = parsed.get(base, {}).get(key, None)
			if val is None:
				# missing entry -> leave blank
				continue
			if abs(val - round(val)) < 1e-9:
				ws.write(start_row + day, j, int(round(val)))
			else:
				ws.write(start_row + day, j, float(val))

	try:
		wb.close()
		print(f"Wrote solution to {filename}")
	except Exception as e:
		print("Could not save Excel file:", e)

#def fetch_params_from